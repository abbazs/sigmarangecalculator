import pathlib

import pandas as pd
from colour import Color
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, Series, StockChart
from openpyxl.chart.axis import ChartLines, DateAxis
from openpyxl.chart.data_source import NumData, NumVal
from openpyxl.chart.label import DataLabel, DataLabelList
from openpyxl.chart.updown_bars import UpDownBars
from openpyxl.styles import Font, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.hyperlink import Hyperlink, HyperlinkList

from log import print_exception
from sigmacols import rangel, rangeu, sigmalmr_cols, sigmaumr_cols


#
def create_line_series(ws, min_col, min_row, max_row, labels, color, legend_loc=0):
    l2 = LineChart()
    l2.add_data(
        Reference(ws, min_col=min_col, min_row=min_row, max_row=max_row),
        titles_from_data=True,
    )
    l2.set_categories(labels)
    l2.series[0].graphicalProperties.line.solidFill = color
    s1 = l2.series[0]
    s1.dLbls = DataLabelList()
    # Initialize data label
    dl = DataLabel()
    # Set properties
    dl.showVal = True
    dl.showSerName = True
    dl.idx = legend_loc
    # position t for top
    dl.position = "r"
    # Append data label to data lebels
    s1.dLbls.dLbl.append(dl)
    return l2


#
def create_work_sheet_chart(ew, df, title, name):
    #
    df.to_excel(excel_writer=ew, sheet_name=name)
    ws = ew.book[name]
    #
    dfl = len(df) + 1
    #
    labels = Reference(ws, min_col=1, min_row=2, max_row=dfl)
    #
    ost = df.columns.get_loc("OPEN") + 2
    cnd = df.columns.get_loc("CLOSE") + 2
    #
    l1 = StockChart()
    data = Reference(ws, min_col=ost, max_col=cnd, min_row=1, max_row=dfl)
    l1.add_data(data, titles_from_data=True)
    l1.set_categories(labels)
    #
    for s in l1.series:
        s.graphicalProperties.line.noFill = True
    #
    l1.hiLowLines = ChartLines()
    l1.upDownBars = UpDownBars()
    if title is not None:
        l1.title = title
    # add dummy cache
    pts = [NumVal(idx=i) for i in range(len(data) - 1)]
    cache = NumData(pt=pts)
    l1.series[-1].val.numRef.numCache = cache
    #
    if dfl <= 6:
        l1.height = 15
        l1.width = 7
    elif dfl >= 6 and dfl <= 25:
        l1.height = 15
        # l1.width = 30
        l1.width = 10
    else:
        l1.height = 20
        # l1.width = 40
        l1.width = 10
    # Monthly constant sigma lines
    clen = max(len(sigmalmr_cols), len(sigmaumr_cols))
    #
    colors = list(Color("#ff4554").range_to(Color("#ffc7cb"), clen))
    colors = [x.get_hex()[1:] for x in colors]
    #
    def create_lines(cols, l1, loc):
        try:
            sli = df.columns.get_loc(cols[0]) + 2
            sln = sli + len(cols)
            for i, xy in enumerate(range(sli, sln)):
                l1 += create_line_series(
                    ws, xy, 1, dfl, labels, colors[i], legend_loc=loc
                )
        except Exception as e:
            print_exception(e)
            print(f"Unable to plot given cols")

    # Creats monthly sigma cols
    create_lines(sigmalmr_cols, l1, 0)
    create_lines(sigmaumr_cols, l1, 0)
    #
    mn = df[sigmalmr_cols[-1]].min() - 100
    mx = df[sigmaumr_cols[-1]].max() + 100
    l1.x_axis.number_format = "yyyymmmdd"
    l1.y_axis.scaling.min = mn
    l1.y_axis.scaling.max = mx
    l1.y_axis.majorUnit = 200
    l1.legend = None
    ws.add_chart(l1, "A2")
    ws.column_dimensions["A"].width = 11
    for cell in ws["A"]:
        cell.style = "custom_datetime"
    ws.column_dimensions["K"].width = 11
    for cell in ws["K"]:
        cell.style = "custom_datetime"
    return ws


#
def create_excel_chart(df, title, name):
    ewb = pd.ExcelWriter(f"{name}.xlsx", engine="openpyxl")
    create_work_sheet_chart(ewb, df, title, 0)
    for name, g in df.groupby("EID"):
        create_work_sheet_chart(ewb, g, None, 1)
    ewb.save()


#
def create_summary_sheet(ew, df, file_name):
    df.to_excel(excel_writer=ew, sheet_name="Summary")
    ws = ew.book["Summary"]
    name = pathlib.Path(file_name).name
    disps = [f"{name}#'{x:%Y-%b-%d}'!A1" for x in df.index]
    cels = [f"A{x}" for x in range(2, len(disps) + 3)]
    for cl, hyp in zip(cels, disps):
        ws[cl].hyperlink = hyp
        ws[cl].hyperlink.location = hyp.split("#")[1]
        ws[cl].style = "custom_datetime"

    cols = ["C", "D", "L"]
    for c in cols:
        for cell in ws[c]:
            cell.number_format = "0.00%"
    #
    ws.column_dimensions["A"].width = 11
    #


#
def add_style(ew):
    ns = NamedStyle(name="custom_datetime", number_format="YYYY-MM-DD")
    ns.font = Font(underline="single", color="0000FF")
    ew.book.add_named_style(ns)


#
def create_summary_percentage_sheet(ew, df):
    df.to_excel(excel_writer=ew, sheet_name="Sigma")
    ws = ew.book["Sigma"]
    ws["A1"].value = "SIGMA"
    #
    for cell in ws["C"]:
        cell.number_format = "0.00%"
    #
    for cell in ws["E"]:
        cell.number_format = "0.00%"
    #
    print("Done summary percentage...")
